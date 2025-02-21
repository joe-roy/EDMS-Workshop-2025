import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

def create_dashboard_workbook(faculty_file, degrees_file, research_file, output_file):
    """
    Create an Excel dashboard from faculty, degrees, and research expenditure data.
    """
    # Read the source data
    faculty_df = pd.read_excel(faculty_file)
    degrees_df = pd.read_excel(degrees_file)
    research_df = pd.read_excel(research_file)
    
    # Calculate total research expenditures
    fund_columns = ['Federal Fund Sum', 'Foreign Fund Sum', 'Individual Fund Sum', 
                   'Industry Fund Sum', 'Local Fund Sum', 'Non-Profit Fund Sum', 
                   'Research Expenditures State Fund Sum']
    research_df['Total Research Expenditures'] = research_df[fund_columns].sum(axis=1)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Create data sheets
    create_data_sheets(wb, faculty_df, degrees_df, research_df)
    
    # Create lookups sheet
    unique_institutions = create_lookups_sheet(wb, faculty_df)
    
    # Create academic metrics dashboard
    create_academic_dashboard(wb, unique_institutions)
    
    # Create research metrics dashboard
    create_research_dashboard(wb, unique_institutions)
    
    # Save the workbook
    wb.save(output_file)

def create_data_sheets(wb, faculty_df, degrees_df, research_df):
    """Create sheets for raw data"""
    # Faculty Data
    faculty_sheet = wb.create_sheet('Faculty Data')
    for r_idx, row in enumerate(faculty_df.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            faculty_sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Degrees Data
    degrees_sheet = wb.create_sheet('Degrees Data')
    for r_idx, row in enumerate(degrees_df.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            degrees_sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Research Data
    research_sheet = wb.create_sheet('Research Data')
    for r_idx, row in enumerate(research_df.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            research_sheet.cell(row=r_idx, column=c_idx, value=value)

def create_lookups_sheet(wb, faculty_df):
    """Create lookups sheet with valid institutions"""
    lookups = wb.create_sheet('Lookups')
    unique_institutions = sorted(faculty_df['School Name'].unique())
    lookups['A1'] = 'Valid Institutions'
    for idx, inst in enumerate(unique_institutions, 2):
        lookups[f'A{idx}'] = inst
    return unique_institutions

def create_academic_dashboard(wb, unique_institutions):
    """Create the academic metrics dashboard"""
    dashboard = wb.create_sheet('Academic Dashboard')
    
    # Header
    dashboard['A1'] = 'Academic Metrics Dashboard'
    dashboard['A2'] = 'Primary Institution:'
    dashboard['A4'] = 'Comparison Institutions:'
    
    # Data validation
    dv = DataValidation(
        type="list",
        formula1=f"=Lookups!$A$2:$A${len(unique_institutions)+1}",
        allow_blank=True
    )
    dashboard.add_data_validation(dv)
    dv.add(dashboard['B2'])
    
    # Comparison rows
    for i in range(5):
        row = i + 5
        dashboard[f'A{row}'] = f'Institution {i+1}'
        dv.add(dashboard[f'B{row}'])
    
    # Metrics section
    metrics_start_row = 12
    dashboard[f'A{metrics_start_row}'] = 'Metrics Summary'
    headers = ['Institution', 'Total Faculty', 'Faculty:Bachelors', 'Faculty:Masters', 'Faculty:Doctoral']
    for col, header in enumerate(headers, 1):
        dashboard.cell(row=metrics_start_row+1, column=col, value=header)
    
    # Add formulas for all institutions
    add_academic_formulas(dashboard, metrics_start_row)
    
    # Add styling
    style_dashboard(dashboard, metrics_start_row, len(headers))
    
    # Add chart
    create_academic_chart(dashboard, metrics_start_row)

def create_research_dashboard(wb, unique_institutions):
    """Create the research metrics dashboard"""
    dashboard = wb.create_sheet('Research Dashboard')
    
    # Header
    dashboard['A1'] = 'Research Metrics Dashboard'
    dashboard['A2'] = 'Primary Institution:'
    dashboard['A4'] = 'Comparison Institutions:'
    
    # Data validation
    dv = DataValidation(
        type="list",
        formula1=f"=Lookups!$A$2:$A${len(unique_institutions)+1}",
        allow_blank=True
    )
    dashboard.add_data_validation(dv)
    dv.add(dashboard['B2'])
    
    # Comparison rows
    for i in range(5):
        row = i + 5
        dashboard[f'A{row}'] = f'Institution {i+1}'
        dv.add(dashboard[f'B{row}'])
    
    # Metrics section
    metrics_start_row = 12
    dashboard[f'A{metrics_start_row}'] = 'Metrics Summary'
    headers = ['Institution', 'Total Faculty', 'Total Research Expenditures', 'Research $ per Faculty']
    for col, header in enumerate(headers, 1):
        dashboard.cell(row=metrics_start_row+1, column=col, value=header)
    
    # Add formulas for all institutions
    add_research_formulas(dashboard, metrics_start_row)
    
    # Add styling
    style_dashboard(dashboard, metrics_start_row, len(headers))
    
    # Add chart
    create_research_chart(dashboard, metrics_start_row)

def add_academic_formulas(dashboard, start_row):
    """Add formulas for academic metrics"""
    # Primary institution
    row = start_row + 2
    dashboard[f'A{row}'] = '=B2'
    
    # Faculty count and ratios
    add_institution_academic_formulas(dashboard, row)
    
    # Comparison institutions
    for i in range(5):
        row = start_row + 3 + i
        dashboard[f'A{row}'] = f'=B{i+5}'
        add_institution_academic_formulas(dashboard, row)

def add_research_formulas(dashboard, start_row):
    """Add formulas for research metrics"""
    # Primary institution
    row = start_row + 2
    dashboard[f'A{row}'] = '=B2'
    
    # Faculty count and research metrics
    add_institution_research_formulas(dashboard, row)
    
    # Comparison institutions
    for i in range(5):
        row = start_row + 3 + i
        dashboard[f'A{row}'] = f'=B{i+5}'
        add_institution_research_formulas(dashboard, row)

def add_institution_academic_formulas(dashboard, row):
    """Add academic formulas for a single institution"""
    # Total Faculty
    dashboard[f'B{row}'] = (
        f'=SUMIFS(\'Faculty Data\'!D:D,\'Faculty Data\'!B:B,$A{row}) + '
        f'SUMIFS(\'Faculty Data\'!E:E,\'Faculty Data\'!B:B,$A{row}) + '
        f'SUMIFS(\'Faculty Data\'!F:F,\'Faculty Data\'!B:B,$A{row})'
    )
    
    # Ratios
    dashboard[f'C{row}'] = f'=IFERROR(B{row}/SUMIFS(\'Degrees Data\'!E:E,\'Degrees Data\'!C:C,$A{row}),0)'
    dashboard[f'D{row}'] = f'=IFERROR(B{row}/SUMIFS(\'Degrees Data\'!F:F,\'Degrees Data\'!C:C,$A{row}),0)'
    dashboard[f'E{row}'] = f'=IFERROR(B{row}/SUMIFS(\'Degrees Data\'!G:G,\'Degrees Data\'!C:C,$A{row}),0)'

def add_institution_research_formulas(dashboard, row):
    """Add research formulas for a single institution"""
    # Total Faculty
    dashboard[f'B{row}'] = (
        f'=SUMIFS(\'Faculty Data\'!D:D,\'Faculty Data\'!B:B,$A{row}) + '
        f'SUMIFS(\'Faculty Data\'!E:E,\'Faculty Data\'!B:B,$A{row}) + '
        f'SUMIFS(\'Faculty Data\'!F:F,\'Faculty Data\'!B:B,$A{row})'
    )
    
    # Total Research Expenditures
    dashboard[f'C{row}'] = (
        f'=SUMIFS(\'Research Data\'!D:D,\'Research Data\'!C:C,$A{row}) + '
        f'SUMIFS(\'Research Data\'!E:E,\'Research Data\'!C:C,$A{row}) + '
        f'SUMIFS(\'Research Data\'!F:F,\'Research Data\'!C:C,$A{row}) + '
        f'SUMIFS(\'Research Data\'!G:G,\'Research Data\'!C:C,$A{row}) + '
        f'SUMIFS(\'Research Data\'!H:H,\'Research Data\'!C:C,$A{row}) + '
        f'SUMIFS(\'Research Data\'!I:I,\'Research Data\'!C:C,$A{row}) + '
        f'SUMIFS(\'Research Data\'!J:J,\'Research Data\'!C:C,$A{row})'
    )
    
    # Research Dollars per Faculty
    dashboard[f'D{row}'] = f'=IFERROR(C{row}/B{row},0)'

def style_dashboard(dashboard, start_row, num_columns):
    """Apply styling to dashboard"""
    dashboard['A1'].font = Font(bold=True, size=14)
    dashboard[f'A{start_row}'].font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row+1, start_row+8):
        for col in range(1, num_columns+1):
            cell = dashboard.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for col in dashboard.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        dashboard.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

def create_academic_chart(dashboard, start_row):
    """Create chart for academic metrics"""
    chart = BarChart()
    chart.title = "Faculty to Degree Ratios by Institution"
    chart.style = 10
    
    for col in range(3, 6):
        data = Reference(dashboard, 
                        min_col=col, 
                        min_row=start_row+1,
                        max_row=start_row+7)
        chart.add_data(data, titles_from_data=True)
    
    cats = Reference(dashboard,
                    min_col=1,
                    min_row=start_row+2,
                    max_row=start_row+7)
    
    chart.set_categories(cats)
    chart.height = 15
    chart.width = 25
    dashboard.add_chart(chart, "G2")

def create_research_chart(dashboard, start_row):
    """Create chart for research metrics"""
    chart = BarChart()
    chart.title = "Research Dollars per Faculty by Institution"
    chart.style = 10
    
    data = Reference(dashboard, 
                    min_col=4,  # Research $ per Faculty column
                    min_row=start_row+1,
                    max_row=start_row+7)
    chart.add_data(data, titles_from_data=True)
    
    cats = Reference(dashboard,
                    min_col=1,
                    min_row=start_row+2,
                    max_row=start_row+7)
    
    chart.set_categories(cats)
    chart.height = 15
    chart.width = 25
    dashboard.add_chart(chart, "G2")

if __name__ == "__main__":
    create_dashboard_workbook(
        faculty_file="Tenure Track Total Faculty.xlsx",
        degrees_file="Total Degrees Awarded.xlsx",
        research_file="Research Expenditures 2023.xlsx",
        output_file="Academic_and_Research_Dashboard.xlsx"
    )